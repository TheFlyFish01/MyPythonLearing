#用途：处理excel文件
#所需模块：
        # openpyxl:处理excel模块
        #Workbook对象属性（工作簿操作）
            #sheetnames：获取工作簿中的表（列表）
            #active：获取当前活跃的Worksheet
            #worksheets：以列表的形式返回所有的Worksheet(表格)
            #read_only：判断是否以read_only模式打开Excel文档
            #encoding：获取文档的字符集编码
            #properties：获取文档的元数据，如标题，创建者，创建日期等
        #Worksheet，Cell对象（工作表操作，单元格）
            #Worksheet:
                #title：表格的标题
                #max_row：表格的最大行
                #min_row：表格的最小行
                #max_column：表格的最大列
                #min_column：表格的最小列
                #rows：按行获取单元格(Cell对象) - 生成器
                #columns：按列获取单元格(Cell对象) - 生成器
                #values：按行获取表格的内容(数据) - 生成器
            #Cell:
                #row：单元格所在的行
                #column：单元格坐在的列
                #value：单元格的值
                #coordinate：单元格的坐标
from openpyxl import Workbook  
from openpyxl import load_workbook 
import string
import datetime 
def create_excel():
    #实例化一个对象
    wb = Workbook()
    #激活worksheet
    ws = wb.active
    #添加数据
    # 方式一：附加行，从第一行开始
    ws.append(['title1','title2'])
    ws.append(['我是title1数据','我是title2数据'])
    # 方式二：数据可以直接分配到单元格中(可以输入公式)
    ws['A1'] = 'title1修改'
    ws['B1'] = 'title2修改'
    # 方式三：Python类型自动转换
    ws['C1'] = '日期'
    ws['C2'] = datetime.datetime.now().strftime('%Y-%m-%d')
    #保存表格
    wb.save('data/test.xlsx')

def open_excel():
    # 打开excel表格
    wb1 = load_workbook('data/test.xlsx')
    # 创建表
        # 方式一：插入到最后(default)
    ws1 = wb1.create_sheet("Mysheet") 
        # 方式二：插入到最开始的位置
    ws2 = wb1.create_sheet("MyFirstsheet", 0)
    #选择表
        #方式一：sheet 名称可以作为 key 进行索引
    sheet3 = wb1["Mysheet"]
    sheet1 = wb1.get_sheet_by_name("MyFirstsheet")
    print(ws1 == sheet3)
        #方式二：获取sheet列表
            # 获取sheets
    sheets = wb1.worksheets
            #打开第二个sheet
    sheet = sheets[1]
    #修改表格数据（假设很多行，可能涉及正则表达式进行匹配）
        # 获取行数
    row_num = sheet.max_row
        # 进行遍历修改（默认标题不修改）
    for i in range(2,row_num+1):
        #获取第一列的值
        data1 = sheet.cell(i,1).value
        #修改第二列的值
        sheet.cell(i,2).value = 'data2'
        print(data1)
    #保存excel表格
    wb1.save('data/test(修改).xlsx')




if __name__ == "__main__":
    #create_excel()
    open_excel()



